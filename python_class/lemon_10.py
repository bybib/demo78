# 读取数据 excel表格
# 文档加载过来   工作簿对象  load_workbooks  最好放在同级路径
import requests,jsonpath
# 导入指定的部分
from openpyxl import load_workbook
# 1.加载工作簿    工作簿名称
# 读取数据的函数
def read_data(filename,sheetname):   # 参数化  文件名  表单名
    wb = load_workbook(filename)
    # print(wb)
    # 2.找表单对象
    sheet = wb[sheetname]
    cases =[]  #空列表用来存用例
    # print(sheet)
    # 3.获取表格对象   获取第一行，第一列
    # cell = sheet.cell(row=1, column=1)
    # 获取单元格的属性
    # print(cell.value)
    # 写入    就是对取得的值进行重新赋值
    # 获取单元哥的属性    ，赋值就是写入   写入生效的话要对工作簿进行保存
    # cell.value = "用例编号"
    # print(cell.value)
    # wb.save("test_case_api.xlsx")     #文件名相同的话是保存，不同的话是另存为

    # 思路：若要读取表单内容，首先确定要读取的内容    url   data  expected
    # 自动获取最大行号
    max_row = sheet.max_row
    for i in range(2, max_row+1):  # 从第二行开始 最大行数是   取头不取尾
        # 把循环的内容放在字典中存储
        case = dict(
            case_id=sheet.cell(row=i, column=1).value,
            url=sheet.cell(row=i, column=5).value,
            # print(url)
            data=sheet.cell(row=i, column=6).value,
            expected=sheet.cell(row=i, column=7).value
        )
        #追加到空列表中
        cases.append(case)
    # print(cases)
    return cases
    # 调用函数
result = read_data("test_case_api.xlsx","register")
# print(result)

#写入函数   可以吧写入的函数也进行函数封装
def write_result(filename,sheetname,final_result,row,column):
    wb = load_workbook(filename)
    # 2.找表单对象
    sheet = wb[sheetname]
    # 找到expected
    cell = sheet.cell(row=row, column=column)
    cell.value = final_result
    # 保存文件
    wb.save(filename)

# 发送几口请求的函数
def api_func(url_api,data_api):
    # 请求头
    head = {"X-Lemonban-Media-Type": "lemonban.v2"}
    response = requests.post(url=url_api,json=data_api,headers=head)
    return response.json()



